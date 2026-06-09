"""
Genera el Excel 'Asignaturas Prioritarias' con análisis comparativo de carga horaria.
"""

import io
from datetime import date
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from analysis import INTL
from data_loader import CURRICULA_LABELS

# ---------------------------------------------------------------------------
# Paleta de colores
# ---------------------------------------------------------------------------
# Encabezados de sección
_HDR_SEC1  = PatternFill("solid", fgColor="C00000")   # rojo oscuro  – exceso
_HDR_SEC2  = PatternFill("solid", fgColor="1F497D")   # azul oscuro  – déficit
_HDR_SEC3  = PatternFill("solid", fgColor="404040")   # gris oscuro  – resumen

# Sub-encabezados de columnas
_SUBHDR1   = PatternFill("solid", fgColor="FFCCCC")   # rojo claro
_SUBHDR2   = PatternFill("solid", fgColor="BDD7EE")   # azul claro
_SUBHDR3   = PatternFill("solid", fgColor="EDEDED")   # gris claro

# Total de sección
_TOTAL_ROW = PatternFill("solid", fgColor="FFF2CC")   # amarillo pálido

# Clasificaciones
_CLASIF_FILL = {
    "EXCLUSIVO SHOA": PatternFill("solid", fgColor="E4C1F9"),  # morado claro
    "CRÍTICO":        PatternFill("solid", fgColor="FF8080"),  # rojo
    "ALTO":           PatternFill("solid", fgColor="FFB347"),  # naranja
    "MEDIO":          PatternFill("solid", fgColor="FFE066"),  # amarillo
    "LEVE":           PatternFill("solid", fgColor="B8E6B0"),  # verde claro
    "DÉFICIT":        PatternFill("solid", fgColor="B0D0F0"),  # azul claro
}
_CLASIF_FONT_COLOR = {
    "EXCLUSIVO SHOA": "5A0080",
    "CRÍTICO":        "7B0000",
    "ALTO":           "6B3000",
    "MEDIO":          "5C4000",
    "LEVE":           "1A4D1A",
    "DÉFICIT":        "0D2E5C",
}

# Bordes
_THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
_MEDIUM_BOTTOM = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="medium"),
)

# Columnas de la tabla
_COLS = [
    "Ranking",
    "Módulo / Asignatura",
    "Horas SHOA",
    "Prom. Internacional",
    "Padilla",
    "Sweden",
    "USS",
    "UCL",
    "Diferencia (h)",
    "Diferencia (%)",
    "Presencia (x/4)",
    "Clasificación",
]
_NC = len(_COLS)   # 12 columnas


# ---------------------------------------------------------------------------
# Funciones auxiliares
# ---------------------------------------------------------------------------

def _presencia(row: pd.Series) -> int:
    """Número de currículos internacionales que tienen ese módulo (>0 horas)."""
    return sum(1 for c in INTL if row.get(c, 0) > 0)


def _classify_sec1(delta_pct: float, presencia: int) -> str:
    """Clasificación para módulos donde SHOA supera al promedio internacional."""
    if presencia == 0:
        return "EXCLUSIVO SHOA"
    if delta_pct > 30 and presencia < 2:
        return "CRÍTICO"
    if delta_pct > 30:
        return "ALTO"
    if delta_pct > 15:
        return "MEDIO"
    return "LEVE"


def _set(ws, row: int, col: int, value=None, *, bold=False, italic=False,
         color="000000", size=10, fill=None, halign="center", wrap=False,
         border=None, num_fmt=None):
    """Escribe una celda con estilos."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, italic=italic, color=color, size=size)
    c.alignment = Alignment(horizontal=halign, vertical="center", wrap_text=wrap)
    if fill:   c.fill   = fill
    if border: c.border = border
    if num_fmt: c.number_format = num_fmt
    return c


def _merge_title(ws, row: int, text: str, fill: PatternFill, font_size=12):
    ws.merge_cells(f"A{row}:{get_column_letter(_NC)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", size=font_size)
    c.fill      = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _write_col_headers(ws, row: int, sub_fill: PatternFill):
    for ci, h in enumerate(_COLS, start=1):
        _set(ws, row, ci, h, bold=True, size=9,
             fill=sub_fill, border=_THIN, halign="center", wrap=True)
    ws.row_dimensions[row].height = 30


def _write_data_section(ws, df_sec: pd.DataFrame, row_start: int,
                         is_deficit: bool = False) -> tuple[int, int, int]:
    """
    Escribe las filas de una sección.
    Returns (next_row_after_total, first_data_row, last_data_row).
    """
    first_row = row_start

    for rank, (_, r) in enumerate(df_sec.iterrows(), start=1):
        pres     = _presencia(r)
        pres_str = f"{pres}/4" if pres > 0 else "0/4 ⚠"
        clasif   = "DÉFICIT" if is_deficit else _classify_sec1(r["delta_avg_pct"], pres)

        vals = [
            rank,
            r["nombre"],
            round(r["shoa"], 1),
            round(r["intl_avg"], 1),
            round(r.get("padilla", 0), 1),
            round(r.get("sweden",  0), 1),
            round(r.get("uss",     0), 1),
            round(r.get("ucl",     0), 1),
            round(r["delta_avg"], 1),
            round(r["delta_avg_pct"], 1),
            pres_str,
            clasif,
        ]

        for ci, val in enumerate(vals, start=1):
            align  = "left" if ci == 2 else "center"
            is_num = isinstance(val, float)
            c = _set(ws, row_start, ci, val,
                     size=9, halign=align, wrap=(ci == 2),
                     border=_THIN,
                     num_fmt="0.0" if is_num and ci != 10 else ("+0.0;-0.0" if is_num else None))

            # Diferencia (h): rojo si positivo, azul si negativo
            if ci == 9 and is_num:
                if val > 0:
                    c.font = Font(bold=True, color="C00000", size=9)
                elif val < 0:
                    c.font = Font(bold=True, color="1F497D", size=9)

            # Diferencia (%): igual
            if ci == 10 and is_num:
                sign_color = "C00000" if val > 0 else "1F497D"
                c.font = Font(bold=True, color=sign_color, size=9)

            # Columna Clasificación: color según tipo
            if ci == 12:
                cfill = _CLASIF_FILL.get(clasif)
                ccolor = _CLASIF_FONT_COLOR.get(clasif, "000000")
                if cfill:   c.fill = cfill
                c.font = Font(bold=True, color=ccolor, size=9)

        row_start += 1

    last_row = row_start - 1

    # --- Fila de totales ---
    for ci in range(1, _NC + 1):
        _set(ws, row_start, ci, border=_MEDIUM_BOTTOM,
             bold=True, size=10, fill=_TOTAL_ROW, halign="center")

    ws.cell(row=row_start, column=1).value = "TOTAL"
    ws.cell(row=row_start, column=2).value = f"{last_row - first_row + 1} módulos"
    ws.cell(row=row_start, column=2).alignment = Alignment(horizontal="left", vertical="center")
    ws.cell(row=row_start, column=2).font = Font(bold=True, size=10)

    for ci in [3, 4, 5, 6, 7, 8, 9]:
        letter = get_column_letter(ci)
        ws.cell(row=row_start, column=ci).value = f"=SUM({letter}{first_row}:{letter}{last_row})"
        ws.cell(row=row_start, column=ci).number_format = "0.0"

    return row_start + 1, first_row, last_row


def _auto_col_widths(ws, min_w=8, max_w=45):
    """Ajusta el ancho de columnas al contenido máximo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_w), max_w)


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def build_priority_excel(df_analyzed: pd.DataFrame) -> bytes:
    """
    Genera el workbook 'Asignaturas Prioritarias'.
    Returns bytes listos para st.download_button.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asignaturas Prioritarias"
    ws.sheet_view.showGridLines = False

    # Separar sección 1 (SHOA > promedio) y sección 2 (SHOA < promedio)
    df_exceso  = df_analyzed[df_analyzed["delta_avg"] > 0].sort_values(
        "delta_avg", ascending=False
    ).reset_index(drop=True)

    df_deficit = df_analyzed[df_analyzed["delta_avg"] < 0].sort_values(
        "delta_avg", ascending=True
    ).reset_index(drop=True)

    # -----------------------------------------------------------------------
    # FILA 1 — Título global
    # -----------------------------------------------------------------------
    cur_row = 1
    ws.merge_cells(f"A1:{get_column_letter(_NC)}1")
    c = ws.cell(row=1, column=1,
                value=f"SHOA — Asignaturas Prioritarias  |  Análisis de Carga Horaria Comparativa  |  {date.today().strftime('%d/%m/%Y')}")
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="1F2D3D")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    cur_row = 2

    # -----------------------------------------------------------------------
    # SECCIÓN 1 — Mayor carga en SHOA
    # -----------------------------------------------------------------------
    ws.row_dimensions[cur_row].height = 6   # espaciado
    cur_row += 1

    _merge_title(ws, cur_row,
                 f"  SECCIÓN 1 — MÓDULOS CON MAYOR CARGA EN SHOA  ({len(df_exceso)} módulos)",
                 _HDR_SEC1)
    cur_row += 1

    sec1_hdr_row = cur_row
    _write_col_headers(ws, cur_row, _SUBHDR1)
    cur_row += 1

    if df_exceso.empty:
        ws.merge_cells(f"A{cur_row}:{get_column_letter(_NC)}{cur_row}")
        ws.cell(row=cur_row, column=1).value = "No hay módulos con mayor carga en SHOA."
        cur_row += 2
        sec1_last_data = cur_row - 2
    else:
        next_row, sec1_first, sec1_last_data = _write_data_section(
            ws, df_exceso, cur_row, is_deficit=False
        )
        cur_row = next_row

        # AutoFilter en sección 1 (único por hoja)
        ws.auto_filter.ref = (
            f"A{sec1_hdr_row}:{get_column_letter(_NC)}{sec1_last_data}"
        )

    # -----------------------------------------------------------------------
    # SECCIÓN 2 — Menor carga en SHOA
    # -----------------------------------------------------------------------
    ws.row_dimensions[cur_row].height = 10  # espaciado entre secciones
    cur_row += 1

    _merge_title(ws, cur_row,
                 f"  SECCIÓN 2 — MÓDULOS CON MENOR CARGA EN SHOA  ({len(df_deficit)} módulos)",
                 _HDR_SEC2)
    cur_row += 1

    _write_col_headers(ws, cur_row, _SUBHDR2)
    cur_row += 1

    if df_deficit.empty:
        ws.merge_cells(f"A{cur_row}:{get_column_letter(_NC)}{cur_row}")
        ws.cell(row=cur_row, column=1).value = "No hay módulos con menor carga en SHOA."
        cur_row += 2
    else:
        next_row, _, _ = _write_data_section(
            ws, df_deficit, cur_row, is_deficit=True
        )
        cur_row = next_row

    # -----------------------------------------------------------------------
    # SECCIÓN 3 — Resumen ejecutivo
    # -----------------------------------------------------------------------
    ws.row_dimensions[cur_row].height = 10
    cur_row += 1

    _merge_title(ws, cur_row,
                 "  SECCIÓN 3 — RESUMEN EJECUTIVO",
                 _HDR_SEC3)
    cur_row += 1

    # Calcular métricas
    total_shoa     = df_analyzed["shoa"].sum()
    total_intl_avg = df_analyzed["intl_avg"].sum()
    diff_total     = total_shoa - total_intl_avg
    n_exceso       = len(df_exceso)
    n_deficit      = len(df_deficit)
    n_exclusivo    = sum(1 for _, r in df_analyzed.iterrows() if _presencia(r) == 0)
    n_critico      = sum(
        1 for _, r in df_exceso.iterrows()
        if _classify_sec1(r["delta_avg_pct"], _presencia(r)) in ("CRÍTICO", "EXCLUSIVO SHOA")
    )

    metricas = [
        ("Total horas SHOA",                   round(total_shoa,     1), "0.0"),
        ("Total horas promedio internacional",  round(total_intl_avg, 1), "0.0"),
        ("Diferencia total (SHOA − Intl)",      round(diff_total,     1), "+0.0;-0.0"),
        ("Módulos con mayor carga en SHOA",     n_exceso,               "0"),
        ("Módulos con menor carga en SHOA",     n_deficit,              "0"),
        ("Módulos exclusivos de SHOA",          n_exclusivo,            "0"),
        ("Módulos CRÍTICO + EXCLUSIVO SHOA",    n_critico,              "0"),
    ]

    # Encabezado tabla resumen
    for ci, h in enumerate(["Métrica", "Valor"], start=1):
        _set(ws, cur_row, ci, h, bold=True, size=10,
             fill=_SUBHDR3, border=_THIN, halign="center")
    cur_row += 1

    for label, val, fmt in metricas:
        _set(ws, cur_row, 1, label, size=10, halign="left",
             border=_THIN, fill=PatternFill("solid", fgColor="F9F9F9"), wrap=True)
        _set(ws, cur_row, 2, val, bold=True, size=10,
             halign="center", border=_THIN, num_fmt=fmt)
        cur_row += 1

    # Top 5 más críticas
    cur_row += 1
    ws.merge_cells(f"A{cur_row}:{get_column_letter(_NC)}{cur_row}")
    c = ws.cell(row=cur_row, column=1, value="  TOP 5 ASIGNATURAS MÁS CRÍTICAS PARA REDUCCIÓN")
    c.font      = Font(bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor="7B2222")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[cur_row].height = 22
    cur_row += 1

    top5_cols = ["Módulo / Asignatura", "Horas SHOA", "Prom. Intl",
                 "Diferencia (h)", "Diferencia (%)", "Presencia", "Clasificación"]
    for ci, h in enumerate(top5_cols, start=1):
        _set(ws, cur_row, ci, h, bold=True, size=9,
             fill=PatternFill("solid", fgColor="E8AAAA"), border=_THIN)
    cur_row += 1

    top5 = df_exceso.head(5)
    for _, r in top5.iterrows():
        pres  = _presencia(r)
        clasif = _classify_sec1(r["delta_avg_pct"], pres)
        cfill  = _CLASIF_FILL.get(clasif)
        ccolor = _CLASIF_FONT_COLOR.get(clasif, "000000")
        top5_vals = [
            r["nombre"],
            round(r["shoa"], 1),
            round(r["intl_avg"], 1),
            round(r["delta_avg"], 1),
            round(r["delta_avg_pct"], 1),
            f"{pres}/4",
            clasif,
        ]
        for ci, val in enumerate(top5_vals, start=1):
            is_num = isinstance(val, float)
            c = _set(ws, cur_row, ci, val, size=9,
                     halign="left" if ci == 1 else "center",
                     border=_THIN, wrap=(ci == 1),
                     num_fmt="0.0" if is_num else None)
            if ci == 7 and cfill:
                c.fill = cfill
                c.font = Font(bold=True, color=ccolor, size=9)
        cur_row += 1

    # -----------------------------------------------------------------------
    # Formato final
    # -----------------------------------------------------------------------
    # Ancho fijo para columna de nombre (col 2), automático para el resto
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    for ci in range(3, _NC + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12

    # Congelar fila de título y primeras 2 columnas
    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
