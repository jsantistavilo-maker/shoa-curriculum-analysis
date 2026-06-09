"""
Módulo de validación independiente del análisis curricular SHOA.

Verifica que los datos procesados por la app coinciden exactamente con el
archivo Excel original, detectando errores de parseo, mapeo de columnas
o celdas mal leídas.

Ejecución directa:
    python validate_data.py
    python validate_data.py --excel          # también genera validacion_datos.xlsx

El módulo es INTENCIONALMENTE independiente de data_loader.py: reimplementa
la lectura cruda para que la comparación sea significativa (no validar una
función contra sí misma).
"""

from __future__ import annotations

import io
import re
import sys
import argparse
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from thefuzz import fuzz


# ---------------------------------------------------------------------------
# Constantes — columnas del Excel (0-indexed, independientes de data_loader)
# ---------------------------------------------------------------------------

# Horas por currículo
_SHOA_H    = [5, 6, 7]
_PADILLA_H = [12, 13, 14]
_SWEDEN_H  = [20, 21, 22, 23]
_USS_H     = [31, 32, 33, 34]
_UCL_H     = [39, 40, 41, 42]

CURRICULA_H = {
    "shoa":    _SHOA_H,
    "padilla": _PADILLA_H,
    "sweden":  _SWEDEN_H,
    "uss":     _USS_H,
    "ucl":     _UCL_H,
}

CURRICULA_LABELS = {
    "shoa":    "SHOA (Chile)",
    "padilla": "Padilla (Colombia)",
    "sweden":  "Sweden",
    "uss":     "USS",
    "ucl":     "UCL",
}

# Columnas de identificación cruzada (para verificar que la fila habla del mismo tópico)
_SHOA_DESC_COL    = 1    # Topic/Element  (inglés)
_PADILLA_DESC_COL = 8    # Topic/Element  (inglés)
_SWEDEN_DESC_COL  = 15   # Topic/Element  (inglés)
_USS_CODE_COL     = 26   # Subcódigo      (debería == código SHOA)
_USS_DESC_COL     = 28   # Contenido      (español — diferente idioma)
_UCL_DESC_COL     = 35   # Topic/Element  (inglés)

UMBRAL_WARN  = 80   # score < 80 → advertencia
UMBRAL_ERROR = 60   # score < 60 → error

# ---------------------------------------------------------------------------
# Utilidades internas
# ---------------------------------------------------------------------------

def _find_file() -> Path | None:
    candidates = [
        Path.home() / "OneDrive" / "Desktop" / "shoa resumen" / "Tabla resumen.xlsx",
        Path.home() / "Desktop"  / "shoa resumen" / "Tabla resumen.xlsx",
        Path.cwd() / "Tabla resumen.xlsx",
        Path(__file__).parent / "Tabla resumen.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return None


def _to_float(val) -> float:
    """Convierte un valor crudo del Excel a float. Implementación independiente."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return 0.0 if (isinstance(val, float) and np.isnan(val)) else max(0.0, float(val))
    try:
        v = str(val).replace(",", ".").strip()
        f = float(v)
        return 0.0 if np.isnan(f) else max(0.0, f)
    except (ValueError, TypeError):
        return 0.0


def _is_leaf(code) -> bool:
    if code is None or (isinstance(code, float) and np.isnan(code)):
        return False
    return bool(re.match(r"^[A-Z]\d+\.\d+[a-z]+$", str(code).strip()))


def _sum_h(row: pd.Series, cols: list[int]) -> float:
    return sum(_to_float(row.iloc[c]) for c in cols if c < len(row))


def _strip_code_prefix(text: str, code: str) -> str:
    """'F1.1a Gravity Field...' → 'Gravity Field...'"""
    t = str(text).strip()
    if t.startswith(code):
        t = t[len(code):].strip()
    return t


# ---------------------------------------------------------------------------
# Lectura cruda del Excel
# ---------------------------------------------------------------------------

def _read_raw() -> tuple[pd.DataFrame | None, str | None]:
    path = _find_file()
    if path is None:
        return None, "Archivo 'Tabla resumen.xlsx' no encontrado."
    try:
        # Lee como bytes primero para evitar bloqueos de Windows cuando
        # el archivo está abierto en Excel o siendo sincronizado por OneDrive.
        data = path.read_bytes()
        df = pd.read_excel(io.BytesIO(data), header=None, sheet_name=0, engine="openpyxl")
        return df, None
    except PermissionError:
        return None, (
            "Acceso denegado al archivo Excel. "
            "Cierra el archivo en Excel (o espera a que OneDrive termine de sincronizar) "
            "y vuelve a ejecutar la validación."
        )
    except Exception as e:
        return None, f"Error al leer el Excel: {e}"


def _extract_raw_leaves(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Extrae todas las filas hoja (códigos tipo F1.1a) directamente del Excel,
    devolviendo valores absolutos por columna de hora.
    """
    rows = []
    for _, row in df_raw.iterrows():
        c0 = row.iloc[0]
        if not _is_leaf(c0):
            continue
        code = str(c0).strip()
        rows.append({
            "codigo": code,
            # Horas absolutas por currículo (sin agrupar)
            "raw_shoa":    _sum_h(row, _SHOA_H),
            "raw_padilla": _sum_h(row, _PADILLA_H),
            "raw_sweden":  _sum_h(row, _SWEDEN_H),
            "raw_uss":     _sum_h(row, _USS_H),
            "raw_ucl":     _sum_h(row, _UCL_H),
            # Descripciones para fuzzy check
            "desc_shoa":    str(row.iloc[_SHOA_DESC_COL]).strip()    if not pd.isna(row.iloc[_SHOA_DESC_COL])    else "",
            "desc_padilla": str(row.iloc[_PADILLA_DESC_COL]).strip() if not pd.isna(row.iloc[_PADILLA_DESC_COL]) else "",
            "desc_sweden":  str(row.iloc[_SWEDEN_DESC_COL]).strip()  if not pd.isna(row.iloc[_SWEDEN_DESC_COL])  else "",
            "uss_code":     str(row.iloc[_USS_CODE_COL]).strip()     if not pd.isna(row.iloc[_USS_CODE_COL])     else "",
            "desc_ucl":     str(row.iloc[_UCL_DESC_COL]).strip()     if not pd.isna(row.iloc[_UCL_DESC_COL])     else "",
        })
    return pd.DataFrame(rows)


# ===========================================================================
# VALIDACIONES
# ===========================================================================

def validate_1_totales(df_raw_leaves: pd.DataFrame, df_processed_leaves: pd.DataFrame) -> dict:
    """
    Validación 1: compara la suma total de horas por currículo.
    Excel crudo vs datos procesados por data_loader.
    """
    results = {}
    for cur, raw_col in [
        ("shoa",    "raw_shoa"),
        ("padilla", "raw_padilla"),
        ("sweden",  "raw_sweden"),
        ("uss",     "raw_uss"),
        ("ucl",     "raw_ucl"),
    ]:
        excel_total = round(df_raw_leaves[raw_col].sum(), 2)
        # En df_processed_leaves la columna se llama igual que el currículo
        app_total   = round(df_processed_leaves[cur].sum(), 2) if cur in df_processed_leaves.columns else None
        diff        = round(abs(excel_total - app_total), 4) if app_total is not None else None
        status      = "OK" if diff == 0 else ("ERROR" if diff > 0 else "N/A")
        results[cur] = {
            "label":       CURRICULA_LABELS[cur],
            "excel_total": excel_total,
            "app_total":   app_total,
            "diferencia":  diff,
            "status":      status,
        }
    return results


def validate_2_conteos(df_raw_leaves: pd.DataFrame, df_processed_leaves: pd.DataFrame) -> dict:
    """
    Validación 2: compara la cantidad de módulos (filas hoja) detectados.
    """
    n_raw  = len(df_raw_leaves)
    n_proc = len(df_processed_leaves)
    diff   = n_raw - n_proc
    lost   = []
    if diff > 0:
        raw_codes  = set(df_raw_leaves["codigo"])
        proc_codes = set(df_processed_leaves["codigo"])
        lost = sorted(raw_codes - proc_codes)
    return {
        "n_excel":       n_raw,
        "n_procesados":  n_proc,
        "diferencia":    diff,
        "codigos_perdidos": lost,
        "status":        "OK" if diff == 0 else "ERROR",
    }


def validate_3_modulo_a_modulo(df_raw_leaves: pd.DataFrame, df_processed_leaves: pd.DataFrame) -> pd.DataFrame:
    """
    Validación 3: compara horas código a código entre Excel crudo y app.
    Detecta cualquier discrepancia numérica.
    """
    proc = df_processed_leaves.set_index("codigo")
    rows = []
    for _, raw_row in df_raw_leaves.iterrows():
        code = raw_row["codigo"]
        proc_row = proc.loc[code] if code in proc.index else None

        for cur, raw_col in [
            ("shoa",    "raw_shoa"),
            ("padilla", "raw_padilla"),
            ("sweden",  "raw_sweden"),
            ("uss",     "raw_uss"),
            ("ucl",     "raw_ucl"),
        ]:
            excel_val = raw_row[raw_col]
            app_val   = float(proc_row[cur]) if proc_row is not None and cur in proc_row else None
            diff      = round(abs(excel_val - app_val), 4) if app_val is not None else None
            rows.append({
                "Código":     code,
                "Currículo":  CURRICULA_LABELS[cur],
                "Excel (h)":  excel_val,
                "App (h)":    app_val,
                "Diferencia": diff,
                "Estado":     "OK" if diff == 0 else ("ERROR" if diff is not None and diff > 0 else "SIN DATOS"),
            })

    return pd.DataFrame(rows)


def validate_4_matching(df_raw_leaves: pd.DataFrame) -> pd.DataFrame:
    """
    Validación 4: coherencia del emparejamiento fila a fila entre currículos.

    Para cada código SHOA verifica:
    - Que el código aparezca en las columnas identificadoras de cada currículo
    - Similitud de nombres (fuzzy) entre SHOA y Padilla / Sweden / UCL
    - Para USS: coincidencia exacta del subcódigo (col 26)

    Nota: el emparejamiento en la app es POSICIONAL (misma fila del Excel),
    no hay fuzzy matching en el pipeline de análisis. Esta validación sirve
    para confirmar que el Excel original tiene los currículos correctamente
    alineados fila a fila.
    """
    rows = []
    for _, r in df_raw_leaves.iterrows():
        code = r["codigo"]

        # ---- Padilla ----
        desc_pad = _strip_code_prefix(r["desc_padilla"], code)
        code_in_padilla = r["desc_padilla"].startswith(code)
        score_pad = fuzz.token_set_ratio(r["desc_shoa"], desc_pad) if desc_pad else 0

        # ---- Sweden ----
        desc_swe = _strip_code_prefix(r["desc_sweden"], code)
        code_in_sweden = r["desc_sweden"].startswith(code)
        score_swe = fuzz.token_set_ratio(r["desc_shoa"], desc_swe) if desc_swe else 0

        # ---- USS: match exacto de código (no nombre — está en español) ----
        uss_code_match = r["uss_code"].strip() == code
        score_uss = 100 if uss_code_match else 0

        # ---- UCL ----
        desc_ucl = _strip_code_prefix(r["desc_ucl"], code)
        code_in_ucl = r["desc_ucl"].startswith(code)
        score_ucl = fuzz.token_set_ratio(r["desc_shoa"], desc_ucl) if desc_ucl else 0

        # Score global (promedio de los 4)
        score_global = round((score_pad + score_swe + score_uss + score_ucl) / 4, 1)

        # Determinar estado
        min_score = min(score_pad, score_swe, score_uss, score_ucl)
        if min_score < UMBRAL_ERROR:
            estado = "ERROR"
        elif min_score < UMBRAL_WARN:
            estado = "REVISAR"
        else:
            estado = "OK"

        rows.append({
            "Código SHOA":         code,
            "Descripción SHOA":    r["desc_shoa"][:50],
            "Padilla — código ✓":  "✓" if code_in_padilla else "✗",
            "Padilla — score":     score_pad,
            "Sweden — código ✓":   "✓" if code_in_sweden  else "✗",
            "Sweden — score":      score_swe,
            "USS — código exacto": "✓" if uss_code_match   else "✗",
            "USS — score":         score_uss,
            "UCL — código ✓":      "✓" if code_in_ucl     else "✗",
            "UCL — score":         score_ucl,
            "Score global":        score_global,
            "Estado":              estado,
        })

    return pd.DataFrame(rows)


def validate_5_casos_especiales(df_raw_leaves: pd.DataFrame) -> dict:
    """
    Validación 5: detecta anomalías y casos especiales en los datos crudos.
    """
    intl_cols = ["raw_padilla", "raw_sweden", "raw_uss", "raw_ucl"]

    # a) Módulos de SHOA con 0 horas pero con horas en algún internacional
    shoa_zero = df_raw_leaves[
        (df_raw_leaves["raw_shoa"] == 0) &
        (df_raw_leaves[intl_cols].sum(axis=1) > 0)
    ][["codigo", "raw_shoa"] + intl_cols].copy()

    # b) Módulos exclusivos de SHOA (0 horas en todos los internacionales)
    exclusivo_shoa = df_raw_leaves[
        (df_raw_leaves["raw_shoa"] > 0) &
        (df_raw_leaves[intl_cols].sum(axis=1) == 0)
    ][["codigo", "raw_shoa"]].copy()

    # c) Valores anómalos: horas muy altas (>200 en un solo elemento)
    all_h_cols = ["raw_shoa"] + intl_cols
    anomalos = []
    for col in all_h_cols:
        cur = col.replace("raw_", "")
        hits = df_raw_leaves[df_raw_leaves[col] > 200][["codigo", col]].copy()
        hits.columns = ["codigo", "horas"]
        hits["curriculo"] = CURRICULA_LABELS.get(cur, cur)
        anomalos.append(hits)
    df_anomalos = pd.concat(anomalos, ignore_index=True) if anomalos else pd.DataFrame()
    if not df_anomalos.empty:
        df_anomalos = df_anomalos[df_anomalos["horas"] > 0].reset_index(drop=True)

    # d) Módulos con horas negativas (nunca debería ocurrir con safe_float)
    negativos = []
    for col in all_h_cols:
        hits = df_raw_leaves[df_raw_leaves[col] < 0][["codigo", col]]
        if not hits.empty:
            negativos.append(hits)
    df_negativos = pd.concat(negativos, ignore_index=True) if negativos else pd.DataFrame()

    return {
        "shoa_zero_intl_positivo": shoa_zero,
        "exclusivo_shoa":          exclusivo_shoa,
        "anomalos":                df_anomalos,
        "negativos":               df_negativos,
    }


# ===========================================================================
# ORQUESTADOR PRINCIPAL
# ===========================================================================

def run_validation() -> dict:
    """
    Ejecuta las 5 validaciones y devuelve un diccionario con todos los resultados.
    Compatible tanto con ejecución directa como con Streamlit.
    """
    # --- Leer Excel crudo (implementación independiente) ---
    df_raw, err = _read_raw()
    if err:
        return {"error": err}

    df_raw_leaves = _extract_raw_leaves(df_raw)

    # --- Cargar datos procesados por data_loader ---
    try:
        from data_loader import load_data
        data, data_err = load_data()
        if data_err:
            return {"error": f"data_loader falló: {data_err}"}
        df_processed_leaves = data["df_leaves"]
    except Exception as e:
        return {"error": f"No se pudo importar data_loader: {e}"}

    # --- Cargar análisis (para casos especiales) ---
    try:
        from analysis import analyze
        df_analyzed = analyze(data["df_subtopics"])
    except Exception as e:
        df_analyzed = None

    # --- Ejecutar validaciones ---
    v1 = validate_1_totales(df_raw_leaves, df_processed_leaves)
    v2 = validate_2_conteos(df_raw_leaves, df_processed_leaves)
    v3 = validate_3_modulo_a_modulo(df_raw_leaves, df_processed_leaves)
    v4 = validate_4_matching(df_raw_leaves)
    v5 = validate_5_casos_especiales(df_raw_leaves)

    # --- Resumen global ---
    errores_v1  = sum(1 for r in v1.values() if r["status"] == "ERROR")
    errores_v3  = len(v3[v3["Estado"] == "ERROR"])
    warns_v4    = len(v4[v4["Estado"].isin(["REVISAR", "ERROR"])]) if not v4.empty else 0
    errores_v4  = len(v4[v4["Estado"] == "ERROR"]) if not v4.empty else 0

    resumen = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "val1": {"status": "OK" if errores_v1 == 0 else "ERROR",
                 "detalle": f"{errores_v1} currículo(s) con diferencia de horas"},
        "val2": {"status": v2["status"],
                 "detalle": f"{v2['n_excel']} en Excel, {v2['n_procesados']} procesados, "
                            f"{v2['diferencia']} diferencia"},
        "val3": {"status": "OK" if errores_v3 == 0 else "ERROR",
                 "detalle": f"{errores_v3} discrepancias numéricas encontradas"},
        "val4": {"status": "OK" if warns_v4 == 0 else ("ERROR" if errores_v4 > 0 else "REVISAR"),
                 "detalle": f"{warns_v4} emparejamientos a revisar, {errores_v4} errores"},
        "val5": {
            "status": "INFO",
            "detalle": (f"{len(v5['exclusivo_shoa'])} módulos exclusivos SHOA · "
                        f"{len(v5['shoa_zero_intl_positivo'])} con SHOA=0 e intl>0 · "
                        f"{len(v5['anomalos'])} valores anómalos (>200h)"),
        },
    }

    return {
        "resumen":     resumen,
        "v1_totales":  v1,
        "v2_conteos":  v2,
        "v3_modulos":  v3,
        "v4_matching": v4,
        "v5_especiales": v5,
        "n_raw_leaves":  len(df_raw_leaves),
    }


# ===========================================================================
# REPORTE DE CONSOLA
# ===========================================================================

def print_report(results: dict):
    """Imprime el reporte de validación en consola con emojis de estado."""
    if "error" in results:
        print(f"❌ ERROR CRÍTICO: {results['error']}")
        return

    ICON = {"OK": "✅", "ERROR": "❌", "REVISAR": "⚠️", "INFO": "ℹ️", "N/A": "—"}

    r = results["resumen"]
    print("\n" + "="*65)
    print(f"  VALIDACIÓN DE DATOS — {r['timestamp']}")
    print("="*65)

    labels_val = [
        ("val1", "Totales por currículo"),
        ("val2", "Cantidad de módulos"),
        ("val3", "Módulo a módulo"),
        ("val4", "Emparejamiento de códigos"),
        ("val5", "Casos especiales"),
    ]
    for key, label in labels_val:
        v = r[key]
        icon = ICON.get(v["status"], "—")
        print(f"  {icon}  {label:<30} {v['detalle']}")

    print()
    print("  — DETALLE VALIDACIÓN 1: Totales por currículo —")
    for cur, v in results["v1_totales"].items():
        icon = ICON.get(v["status"], "—")
        print(f"  {icon}  {v['label']:<30} "
              f"Excel={v['excel_total']:>8.1f}h  "
              f"App={v['app_total']:>8.1f}h  "
              f"Δ={v['diferencia']:>6.2f}")

    print()
    print("  — DETALLE VALIDACIÓN 2: Conteo de módulos —")
    v2 = results["v2_conteos"]
    print(f"  {'✅' if v2['status']=='OK' else '❌'}  "
          f"Excel: {v2['n_excel']} filas  |  "
          f"App: {v2['n_procesados']} filas  |  "
          f"Δ={v2['diferencia']}")
    if v2["codigos_perdidos"]:
        print(f"     Códigos no procesados: {', '.join(v2['codigos_perdidos'][:10])}"
              f"{'...' if len(v2['codigos_perdidos'])>10 else ''}")

    print()
    print("  — DETALLE VALIDACIÓN 3: Discrepancias numéricas —")
    v3 = results["v3_modulos"]
    errores = v3[v3["Estado"] == "ERROR"]
    if errores.empty:
        print("  ✅  Sin discrepancias numéricas.")
    else:
        print(f"  ❌  {len(errores)} discrepancias:")
        for _, row in errores.head(10).iterrows():
            print(f"     {row['Código']:<10} {row['Currículo']:<22} "
                  f"Excel={row['Excel (h)']:>7.1f}  App={row['App (h)']:>7.1f}  "
                  f"Δ={row['Diferencia']:>6.2f}")

    print()
    print("  — DETALLE VALIDACIÓN 4: Emparejamiento de códigos —")
    v4 = results["v4_matching"]
    if v4.empty:
        print("  —  Sin datos de emparejamiento.")
    else:
        warn_rows  = v4[v4["Estado"].isin(["REVISAR", "ERROR"])]
        error_rows = v4[v4["Estado"] == "ERROR"]
        print(f"  {'✅' if len(warn_rows)==0 else '⚠️ '}  "
              f"{len(v4)} códigos verificados — "
              f"{len(warn_rows)} a revisar (score<{UMBRAL_WARN}), "
              f"{len(error_rows)} errores (score<{UMBRAL_ERROR})")
        if not warn_rows.empty:
            print("     Primeros casos a revisar:")
            for _, row in warn_rows.head(5).iterrows():
                print(f"     {row['Código SHOA']:<10}  "
                      f"Pad:{row['Padilla — score']:>3}  "
                      f"Swe:{row['Sweden — score']:>3}  "
                      f"USS:{row['USS — score']:>3}  "
                      f"UCL:{row['UCL — score']:>3}  "
                      f"→ {row['Estado']}")

    print()
    print("  — DETALLE VALIDACIÓN 5: Casos especiales —")
    v5 = results["v5_especiales"]
    print(f"  ℹ️   Módulos exclusivos SHOA:          {len(v5['exclusivo_shoa'])}")
    print(f"  ℹ️   SHOA=0h, internacionales>0h:      {len(v5['shoa_zero_intl_positivo'])}")
    print(f"  ℹ️   Valores anómalos (>200h):          {len(v5['anomalos'])}")
    print(f"  {'✅' if v5['negativos'].empty else '❌'}   Horas negativas detectadas:       "
          f"{len(v5['negativos'])}")
    print("="*65)


# ===========================================================================
# EXPORTACIÓN A EXCEL
# ===========================================================================

# Estilos reutilizables
_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_FILLS = {
    "OK":     PatternFill("solid", fgColor="C6EFCE"),   # verde
    "ERROR":  PatternFill("solid", fgColor="FFC7CE"),   # rojo
    "REVISAR":PatternFill("solid", fgColor="FFEB9C"),   # amarillo
    "INFO":   PatternFill("solid", fgColor="BDD7EE"),   # azul
    "N/A":    PatternFill("solid", fgColor="EDEDED"),   # gris
}
_HDR_FILL  = PatternFill("solid", fgColor="1F3864")
_HDR2_FILL = PatternFill("solid", fgColor="2F75B6")


def _hdr(ws, row: int, cols: list[str], fill=None):
    fill = fill or _HDR2_FILL
    for ci, label in enumerate(cols, start=1):
        c = ws.cell(row=row, column=ci, value=label)
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.fill      = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws.row_dimensions[row].height = 28


def _row(ws, row_idx: int, values: list, fills: list | None = None):
    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.alignment = Alignment(
            horizontal="left" if isinstance(val, str) else "center",
            vertical="center",
            wrap_text=False,
        )
        c.border = _THIN
        c.font   = Font(size=9)
        if fills and ci - 1 < len(fills) and fills[ci - 1]:
            c.fill = fills[ci - 1]


def _auto_widths(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, max_w)


def _sheet_title(ws, title: str, n_cols: int):
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(bold=True, size=13, color="FFFFFF")
    c.fill      = _HDR_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26


def build_validation_excel(results: dict) -> bytes:
    """Genera el workbook de validación con 5 hojas."""
    if "error" in results:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = f"ERROR: {results['error']}"
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # quitar hoja vacía

    # ------------------------------------------------------------------
    # HOJA 1 — Resumen general (semáforo)
    # ------------------------------------------------------------------
    ws1 = wb.create_sheet("1 Resumen General")
    _sheet_title(ws1, f"Resumen de Validación  —  {results['resumen']['timestamp']}", 4)
    _hdr(ws1, 2, ["N°", "Validación", "Estado", "Detalle"])

    val_labels = [
        ("val1", "Totales por currículo"),
        ("val2", "Cantidad de módulos"),
        ("val3", "Módulo a módulo (numérico)"),
        ("val4", "Emparejamiento de códigos"),
        ("val5", "Casos especiales y anomalías"),
    ]
    for i, (key, label) in enumerate(val_labels, start=1):
        v = results["resumen"][key]
        estado_fill = _FILLS.get(v["status"], _FILLS["N/A"])
        _row(ws1, i + 2, [i, label, v["status"], v["detalle"]],
             fills=[None, None, estado_fill, None])

    ws1.column_dimensions["A"].width = 5
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 55

    # ------------------------------------------------------------------
    # HOJA 2 — Totales por currículo
    # ------------------------------------------------------------------
    ws2 = wb.create_sheet("2 Totales por Curriculo")
    _sheet_title(ws2, "Validación 1 — Totales de horas por currículo", 5)
    _hdr(ws2, 2, ["Currículo", "Horas Excel (crudo)", "Horas App (procesado)", "Diferencia", "Estado"])

    for i, (cur, v) in enumerate(results["v1_totales"].items(), start=3):
        fill = _FILLS.get(v["status"])
        _row(ws2, i, [
            v["label"],
            v["excel_total"],
            v["app_total"],
            v["diferencia"],
            v["status"],
        ], fills=[None, None, None, fill, fill])
        for ci in [2, 3, 4]:
            ws2.cell(row=i, column=ci).number_format = "0.00"

    _auto_widths(ws2)

    # ------------------------------------------------------------------
    # HOJA 3 — Módulo a módulo
    # ------------------------------------------------------------------
    ws3 = wb.create_sheet("3 Modulo a Modulo")
    v3  = results["v3_modulos"]
    _sheet_title(ws3, "Validación 3 — Comparación numérica código a código", len(v3.columns))
    _hdr(ws3, 2, list(v3.columns))

    ws3.auto_filter.ref = f"A2:{get_column_letter(len(v3.columns))}2"

    for r_idx, (_, row) in enumerate(v3.iterrows(), start=3):
        estado = str(row.get("Estado", ""))
        fill   = _FILLS.get(estado)
        values = [row[c] for c in v3.columns]
        fills  = [fill if c == "Estado" else None for c in v3.columns]
        _row(ws3, r_idx, values, fills=fills)

    _auto_widths(ws3)

    # ------------------------------------------------------------------
    # HOJA 4 — Emparejamiento de códigos (fuzzy)
    # ------------------------------------------------------------------
    ws4 = wb.create_sheet("4 Matching de Codigos")
    v4  = results["v4_matching"]

    if v4.empty:
        ws4["A1"] = "Sin datos de emparejamiento."
    else:
        _sheet_title(ws4, "Validación 4 — Coherencia del emparejamiento fila a fila entre currículos", len(v4.columns))
        # Nota explicativa
        ws4.merge_cells(f"A2:{get_column_letter(len(v4.columns))}2")
        c = ws4.cell(row=2, column=1,
                     value="Nota: el emparejamiento es POSICIONAL (misma fila del Excel). "
                           "Los scores de Padilla/Sweden/UCL miden similitud de nombres (fuzz.token_set_ratio). "
                           "USS se valida por coincidencia exacta de código.")
        c.font = Font(italic=True, size=9, color="444444")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws4.row_dimensions[2].height = 30

        _hdr(ws4, 3, list(v4.columns))
        ws4.auto_filter.ref = f"A3:{get_column_letter(len(v4.columns))}3"

        for r_idx, (_, row) in enumerate(v4.iterrows(), start=4):
            estado = str(row.get("Estado", ""))
            values = [row[c] for c in v4.columns]
            fills  = []
            for c_name in v4.columns:
                if c_name == "Estado":
                    fills.append(_FILLS.get(estado))
                elif "score" in c_name.lower():
                    score = row[c_name]
                    if isinstance(score, (int, float)):
                        if score < UMBRAL_ERROR:
                            fills.append(_FILLS["ERROR"])
                        elif score < UMBRAL_WARN:
                            fills.append(_FILLS["REVISAR"])
                        else:
                            fills.append(_FILLS["OK"])
                    else:
                        fills.append(None)
                else:
                    fills.append(None)
            _row(ws4, r_idx, values, fills=fills)

    _auto_widths(ws4)

    # ------------------------------------------------------------------
    # HOJA 5 — Casos especiales
    # ------------------------------------------------------------------
    ws5 = wb.create_sheet("5 Casos Especiales")
    _sheet_title(ws5, "Validación 5 — Casos especiales y anomalías", 7)
    cur_row = 3

    def _sec(title: str, fill_color: str):
        nonlocal cur_row
        ws5.merge_cells(f"A{cur_row}:G{cur_row}")
        c = ws5.cell(row=cur_row, column=1, value=f"  {title}")
        c.font      = Font(bold=True, color="FFFFFF", size=11)
        c.fill      = PatternFill("solid", fgColor=fill_color)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws5.row_dimensions[cur_row].height = 22
        cur_row += 1

    v5 = results["v5_especiales"]

    # a) SHOA = 0, internacionales > 0
    _sec("a) Módulos con SHOA=0h pero con horas en currículos internacionales", "2F75B6")
    df_a = v5["shoa_zero_intl_positivo"]
    if df_a.empty:
        ws5.cell(row=cur_row, column=1).value = "Ninguno encontrado."
        cur_row += 2
    else:
        cols_a = list(df_a.columns)
        _hdr(ws5, cur_row, cols_a, fill=PatternFill("solid", fgColor="9DC3E6"))
        cur_row += 1
        for _, row in df_a.iterrows():
            _row(ws5, cur_row, [row[c] for c in cols_a])
            for ci in range(2, len(cols_a) + 1):
                ws5.cell(row=cur_row, column=ci).number_format = "0.0"
            cur_row += 1
        cur_row += 1

    # b) Módulos exclusivos de SHOA
    _sec("b) Módulos exclusivos de SHOA (0h en todos los currículos internacionales)", "7030A0")
    df_b = v5["exclusivo_shoa"]
    if df_b.empty:
        ws5.cell(row=cur_row, column=1).value = "Ninguno encontrado."
        cur_row += 2
    else:
        _hdr(ws5, cur_row, ["Código", "Horas SHOA"],
             fill=PatternFill("solid", fgColor="D9B3FF"))
        cur_row += 1
        for _, row in df_b.iterrows():
            _row(ws5, cur_row, [row["codigo"], row["raw_shoa"]])
            ws5.cell(row=cur_row, column=2).number_format = "0.0"
            cur_row += 1
        cur_row += 1

    # c) Valores anómalos
    _sec("c) Valores anómalos (>200h en un único elemento)", "C55A11")
    df_c = v5["anomalos"]
    if df_c.empty:
        ws5.cell(row=cur_row, column=1).value = "Ninguno encontrado."
        cur_row += 2
    else:
        _hdr(ws5, cur_row, ["Código", "Currículo", "Horas"],
             fill=PatternFill("solid", fgColor="F4B183"))
        cur_row += 1
        for _, row in df_c.iterrows():
            _row(ws5, cur_row, [row["codigo"], row["curriculo"], row["horas"]])
            ws5.cell(row=cur_row, column=3).number_format = "0.0"
            cur_row += 1
        cur_row += 1

    # d) Valores negativos
    _sec("d) Horas negativas (no debería ocurrir)", "C00000")
    df_d = v5["negativos"]
    if df_d.empty:
        ws5.cell(row=cur_row, column=1).value = "✅ Ningún valor negativo detectado."
    else:
        for _, row in df_d.iterrows():
            _row(ws5, cur_row, list(row))
            cur_row += 1

    _auto_widths(ws5)

    # ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===========================================================================
# PUNTO DE ENTRADA DIRECTO
# ===========================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Valida los datos procesados contra el Excel original."
    )
    parser.add_argument(
        "--excel", action="store_true",
        help="Genera también el archivo validacion_datos.xlsx"
    )
    args = parser.parse_args()

    print("Ejecutando validación de datos…")
    results = run_validation()
    print_report(results)

    if args.excel or "--excel" in sys.argv:
        out_path = Path.cwd() / "validacion_datos.xlsx"
        xlsx = build_validation_excel(results)
        out_path.write_bytes(xlsx)
        print(f"\n📄 Excel guardado en: {out_path}")


if __name__ == "__main__":
    main()
