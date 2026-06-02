"""
Análisis curricular por asignatura.

Agrupa las filas hoja por nombre de asignatura SHOA (columna E) y suma las
horas de todos los currículos en el mismo contexto temático.
Clasifica cada asignatura como SOBREESTIMADA / SUBESTIMADA / ALINEADA /
EXCLUSIVA SHOA y genera el Excel "Carga por Asignatura".
"""

from __future__ import annotations

import io
from datetime import date

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INTL_KEYS  = ["padilla", "sweden", "uss", "ucl"]
THRESHOLD  = 20.0   # ±20 % → SOBREESTIMADA / SUBESTIMADA

# ── Colores Excel ──────────────────────────────────────────────────────────
_H    = PatternFill("solid", fgColor="003366")   # encabezado azul marino
_S1   = PatternFill("solid", fgColor="003366")   # sección 1
_S2   = PatternFill("solid", fgColor="1F497D")   # sección 2
_S3   = PatternFill("solid", fgColor="404040")   # sección 3
_ALT  = PatternFill("solid", fgColor="F5F5F5")   # fila alternada
_TOT  = PatternFill("solid", fgColor="FFF2CC")   # fila total

CLASIF_FILL = {
    "SOBREESTIMADA":  PatternFill("solid", fgColor="FADBD8"),
    "SUBESTIMADA":    PatternFill("solid", fgColor="D6EAF8"),
    "ALINEADA":       PatternFill("solid", fgColor="D5F5E3"),
    "EXCLUSIVA SHOA": PatternFill("solid", fgColor="E8DAEF"),
}

CLASIF_COLORS_HEX = {
    "SOBREESTIMADA":  "#C0392B",
    "SUBESTIMADA":    "#2471A3",
    "ALINEADA":       "#1A7A4A",
    "EXCLUSIVA SHOA": "#6C3483",
}

_THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)


# ── Helpers internos ───────────────────────────────────────────────────────

def _mode_name(series: pd.Series) -> str:
    """Valor más frecuente no nulo de una serie de strings."""
    vals = series.dropna().astype(str)
    vals = vals[vals.str.strip().ne("") & ~vals.str.lower().isin({"nan", "none"})]
    if vals.empty:
        return "—"
    m = vals.mode()
    return m.iloc[0] if not m.empty else "—"


def _delta_pct(shoa: float, avg: float) -> float:
    if avg > 0:
        return (shoa - avg) / avg * 100
    return 100.0 if shoa > 0 else 0.0


def _classify(shoa: float, intl_vals: list[float]) -> str:
    if all(h == 0 for h in intl_vals) and shoa > 0:
        return "EXCLUSIVA SHOA"
    avg = float(np.mean(intl_vals))
    pct = _delta_pct(shoa, avg)
    if pct > THRESHOLD:
        return "SOBREESTIMADA"
    if pct < -THRESHOLD:
        return "SUBESTIMADA"
    return "ALINEADA"


# ── Análisis principal ─────────────────────────────────────────────────────

def compute_assignment_analysis(df_leaves: pd.DataFrame) -> pd.DataFrame:
    """
    Agrupa df_leaves por asgn_shoa y retorna un DataFrame con una fila por
    asignatura SHOA, con horas propias e internacionales equivalentes.
    """
    if "asgn_shoa" not in df_leaves.columns:
        return pd.DataFrame()

    df = df_leaves[df_leaves["asgn_shoa"].notna()].copy()
    if df.empty:
        return pd.DataFrame()

    rows = []
    for asgn, grp in df.groupby("asgn_shoa", sort=False):
        shoa_h = grp["shoa"].sum()
        ph = grp["padilla"].sum()
        sh = grp["sweden"].sum()
        uh = grp["uss"].sum()
        ch = grp["ucl"].sum()

        intl_vals = [ph, sh, uh, ch]
        avg = float(np.mean(intl_vals))
        dh  = shoa_h - avg
        dp  = _delta_pct(shoa_h, avg)
        clf = _classify(shoa_h, intl_vals)

        subtopics = sorted(grp["subtopico"].unique().tolist())

        rows.append({
            "asignatura_shoa": asgn,
            "horas_shoa":      round(shoa_h, 1),
            "horas_padilla":   round(ph, 1),
            "horas_sweden":    round(sh, 1),
            "horas_uss":       round(uh, 1),
            "horas_ucl":       round(ch, 1),
            "asgn_padilla":    _mode_name(grp["asgn_padilla"]) if "asgn_padilla" in grp else "—",
            "asgn_sweden":     _mode_name(grp["asgn_sweden"])  if "asgn_sweden"  in grp else "—",
            "asgn_uss":        _mode_name(grp["asgn_uss"])     if "asgn_uss"     in grp else "—",
            "asgn_ucl":        _mode_name(grp["asgn_ucl"])     if "asgn_ucl"     in grp else "—",
            "intl_avg":        round(avg, 1),
            "delta_h":         round(dh,  1),
            "delta_pct":       round(dp,  1),
            "delta_padilla":   round(shoa_h - ph, 1),
            "delta_sweden":    round(shoa_h - sh, 1),
            "delta_uss":       round(shoa_h - uh, 1),
            "delta_ucl":       round(shoa_h - ch, 1),
            "clasificacion":   clf,
            "n_subtopicos":    len(subtopics),
            "subtopicos":      "; ".join(subtopics),
            "pct_total":       0.0,
        })

    df_res = (
        pd.DataFrame(rows)
        .sort_values("horas_shoa", ascending=False)
        .reset_index(drop=True)
    )

    total = df_res["horas_shoa"].sum()
    if total > 0:
        df_res["pct_total"] = (df_res["horas_shoa"] / total * 100).round(1)

    return df_res


def compute_assignment_kpis(df_a: pd.DataFrame) -> dict:
    if df_a.empty:
        return {"n_total": 0, "n_sobre": 0, "n_sub": 0, "n_alin": 0, "n_excl": 0,
                "top_sobre": None, "top_sub": None}

    counts = df_a["clasificacion"].value_counts().to_dict()

    top_sobre = None
    df_s = df_a[df_a["clasificacion"] == "SOBREESTIMADA"]
    if not df_s.empty:
        r = df_s.nlargest(1, "delta_h").iloc[0]
        top_sobre = {"nombre": r["asignatura_shoa"], "delta": float(r["delta_h"])}

    top_sub = None
    df_u = df_a[df_a["clasificacion"] == "SUBESTIMADA"]
    if not df_u.empty:
        r = df_u.nsmallest(1, "delta_h").iloc[0]
        top_sub = {"nombre": r["asignatura_shoa"], "delta": float(r["delta_h"])}

    return {
        "n_total": len(df_a),
        "n_sobre": counts.get("SOBREESTIMADA",  0),
        "n_sub":   counts.get("SUBESTIMADA",    0),
        "n_alin":  counts.get("ALINEADA",       0),
        "n_excl":  counts.get("EXCLUSIVA SHOA", 0),
        "top_sobre": top_sobre,
        "top_sub":   top_sub,
    }


# ── Helpers Excel ──────────────────────────────────────────────────────────

def _cell(ws, row, col, val=None, *, bold=False, color="000000", size=10,
          fill=None, ha="center", wrap=False, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fill:   c.fill   = fill
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c


def _sec_title(ws, row, text, fill, nc):
    ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color="FFFFFF", size=12)
    c.fill      = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _hdr_row(ws, row, cols):
    for ci, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill      = _H
        c.font      = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN
    ws.row_dimensions[row].height = 30


def _auto_widths(ws, overrides: dict | None = None):
    for col_cells in ws.columns:
        letter  = get_column_letter(col_cells[0].column)
        if overrides and letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
            continue
        max_len = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=8,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 3, 8), 45)


# ── Excel principal ────────────────────────────────────────────────────────

def build_assignment_excel(df_a: pd.DataFrame) -> bytes:
    """
    Genera el workbook con la hoja 'Carga por Asignatura' (3 secciones).
    Retorna bytes listos para st.download_button.
    """
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Carga por Asignatura"
    ws.sheet_view.showGridLines = False

    if df_a.empty:
        ws["A1"] = "No se encontraron datos de asignaturas en el archivo."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    kpis = compute_assignment_kpis(df_a)
    NC   = 14    # columnas máximas (sección 2)
    cr   = 1

    # ── Título global ─────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    c = ws.cell(row=1, column=1,
                value=f"SHOA — Análisis de Carga por Asignatura  |  {date.today().strftime('%d/%m/%Y')}")
    c.font      = Font(bold=True, size=14, color="FFFFFF")
    c.fill      = _S1
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    cr = 3

    # ──────────────────────────────────────────────────────────────────────────
    # SECCIÓN 1 — Carga total por asignatura SHOA
    # ──────────────────────────────────────────────────────────────────────────
    _sec_title(ws, cr, f"  SECCIÓN 1 — SHOA: Carga Total por Asignatura  ({len(df_a)} asignaturas)",
               _S1, NC)
    cr += 1

    S1 = ["Asignatura SHOA", "Total Horas SHOA", "N° Tópicos que Cubre",
          "Tópicos Asociados (lista)", "% del Total del Currículo"]
    _hdr_row(ws, cr, S1)
    s1_hdr = cr
    cr += 1

    for i, (_, r) in enumerate(df_a.iterrows()):
        alt = _ALT if i % 2 else None
        data = [
            (r["asignatura_shoa"], None,   "left",   True),
            (r["horas_shoa"],      "0.0",  "center", False),
            (r["n_subtopicos"],    None,   "center", False),
            (r["subtopicos"],      None,   "left",   True),
            (r["pct_total"],       "0.0",  "center", False),
        ]
        for ci, (val, fmt, ha, wrap) in enumerate(data, start=1):
            c = ws.cell(row=cr, column=ci, value=val)
            c.border    = _THIN
            c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
            if alt: c.fill = alt
            if fmt: c.number_format = fmt
        cr += 1

    # Fila total
    for ci in range(1, 6):
        ws.cell(row=cr, column=ci).fill   = _TOT
        ws.cell(row=cr, column=ci).border = _THIN
    ws.cell(row=cr, column=1, value="TOTAL").font      = Font(bold=True, size=10)
    ws.cell(row=cr, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=cr, column=2, value=round(df_a["horas_shoa"].sum(), 1)).font = Font(bold=True, size=10)
    ws.cell(row=cr, column=2).number_format = "0.0"
    ws.cell(row=cr, column=2).alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A{s1_hdr}:E{cr - 1}"
    cr += 2

    # ──────────────────────────────────────────────────────────────────────────
    # SECCIÓN 2 — Comparativa de asignaturas equivalentes
    # ──────────────────────────────────────────────────────────────────────────
    _sec_title(ws, cr, "  SECCIÓN 2 — Comparativa de Asignaturas Equivalentes", _S2, NC)
    cr += 1

    S2 = [
        "Asignatura SHOA", "Horas SHOA",
        "Asignatura Padilla", "Horas Padilla",
        "Asignatura SWD", "Horas SWD",
        "Asignatura USS", "Horas USS",
        "Asignatura UCL", "Horas UCL",
        "Promedio Internacional", "Diferencia (h)", "Diferencia (%)", "Clasificación",
    ]
    _hdr_row(ws, cr, S2)
    cr += 1

    for _, r in df_a.iterrows():
        clf  = r["clasificacion"]
        rfil = CLASIF_FILL.get(clf)
        row_data = [
            (r["asignatura_shoa"],  None,        "left",   True),
            (r["horas_shoa"],       "0.0",       "center", False),
            (r["asgn_padilla"],     None,        "left",   True),
            (r["horas_padilla"],    "0.0",       "center", False),
            (r["asgn_sweden"],      None,        "left",   True),
            (r["horas_sweden"],     "0.0",       "center", False),
            (r["asgn_uss"],         None,        "left",   True),
            (r["horas_uss"],        "0.0",       "center", False),
            (r["asgn_ucl"],         None,        "left",   True),
            (r["horas_ucl"],        "0.0",       "center", False),
            (r["intl_avg"],         "0.0",       "center", False),
            (r["delta_h"],          "+0.0;-0.0", "center", False),
            (r["delta_pct"],        "+0.0;-0.0", "center", False),
            (clf,                   None,        "center", False),
        ]
        for ci, (val, fmt, ha, wrap) in enumerate(row_data, start=1):
            c = ws.cell(row=cr, column=ci, value=val)
            c.border    = _THIN
            c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
            if rfil: c.fill = rfil
            if fmt:  c.number_format = fmt
            if ci == 13 and isinstance(val, (int, float)):
                c.font = Font(bold=True,
                              color="CC0000" if val > 0 else "003399",
                              size=10)
        cr += 1

    cr += 2

    # ──────────────────────────────────────────────────────────────────────────
    # SECCIÓN 3 — Resumen ejecutivo
    # ──────────────────────────────────────────────────────────────────────────
    _sec_title(ws, cr, "  SECCIÓN 3 — Resumen Ejecutivo", _S3, NC)
    cr += 1

    _hdr_row(ws, cr, ["Métrica", "Valor"])
    cr += 1

    total = kpis["n_total"]
    pct   = lambda n: f"{round(n / total * 100, 1)}%" if total else "0%"

    metricas = [
        ("Total asignaturas SHOA",           total),
        ("Asignaturas sobreestimadas",        f"{kpis['n_sobre']}  ({pct(kpis['n_sobre'])})"),
        ("Asignaturas subestimadas",          f"{kpis['n_sub']}  ({pct(kpis['n_sub'])})"),
        ("Asignaturas alineadas",             f"{kpis['n_alin']}  ({pct(kpis['n_alin'])})"),
        ("Asignaturas exclusivas de SHOA",    f"{kpis['n_excl']}  ({pct(kpis['n_excl'])})"),
        ("Mayor sobreestimación",
         f"{kpis['top_sobre']['nombre']}  (+{kpis['top_sobre']['delta']:.1f} h)"
         if kpis["top_sobre"] else "—"),
        ("Mayor subestimación",
         f"{kpis['top_sub']['nombre']}  ({kpis['top_sub']['delta']:.1f} h)"
         if kpis["top_sub"] else "—"),
    ]

    for i, (lbl, val) in enumerate(metricas):
        alt = _ALT if i % 2 else None
        c1 = ws.cell(row=cr, column=1, value=lbl)
        c1.border    = _THIN
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if alt: c1.fill = alt

        c2 = ws.cell(row=cr, column=2, value=val)
        c2.border    = _THIN
        c2.font      = Font(bold=True, size=10)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        if alt: c2.fill = alt
        cr += 1

    # ── Anchos de columna ─────────────────────────────────────────────────────
    _auto_widths(ws, overrides={
        "A": 38, "B": 13, "C": 32, "D": 13,
        "E": 32, "F": 13, "G": 32, "H": 13,
        "I": 32, "J": 13, "K": 14, "L": 14,
        "M": 14, "N": 18,
    })
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
