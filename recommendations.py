"""
Módulo de recomendaciones curriculares.
Genera tablas de recomendaciones y exporta a Excel con formato.
"""

from __future__ import annotations

import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from analysis import INTL, suggested_hours, best_reference
from data_loader import CURRICULA_LABELS

# Colores de urgencia para el Excel exportado
FILL_ALTA   = PatternFill("solid", fgColor="FF4444")   # rojo
FILL_MEDIA  = PatternFill("solid", fgColor="FFAA00")   # naranja
FILL_BAJO   = PatternFill("solid", fgColor="44BB44")   # verde
FILL_HEADER = PatternFill("solid", fgColor="003366")   # azul marino (encabezados)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def build_recommendations(df_analyzed: pd.DataFrame) -> pd.DataFrame:
    """
    Genera el DataFrame de recomendaciones curriculares.

    Columnas:
        Módulo | Clasificación | Horas SHOA | Promedio Internacional |
        Diferencia | % Diferencia | Urgencia | Horas Sugeridas |
        Reducción/Aumento | Referencia Internacional
    """
    rows = []
    for _, row in df_analyzed.iterrows():
        clasif = row["clasificacion"]
        if clasif == "ALINEADO":
            continue  # solo módulos desalineados en la tabla

        sugeridas = suggested_hours(row)
        diferencia = row["shoa"] - sugeridas
        referencia = best_reference(row)

        # Qué curriculos tienen MENOS horas que SHOA (para sobrevalorados)
        curricula_menores = [
            CURRICULA_LABELS.get(c, c) for c in INTL if row["shoa"] > row[c]
        ]
        curricula_mayores = [
            CURRICULA_LABELS.get(c, c) for c in INTL if row[c] > row["shoa"]
        ]

        rows.append({
            "Módulo":                   row["nombre"],
            "Clasificación":            clasif,
            "Horas SHOA":               round(row["shoa"], 1),
            "Promedio Internacional":   round(row["intl_avg"], 1),
            "Diferencia (h)":           round(row["delta_avg"], 1),
            "% Diferencia":             round(row["delta_avg_pct"], 1),
            "Urgencia":                 row["urgencia"],
            "Horas Sugeridas":          sugeridas,
            "Ajuste (h)":               round(-diferencia, 1),   # positivo = aumentar
            "Referencia Internacional": referencia,
            "Curricula con menor carga": "; ".join(curricula_menores) if curricula_menores else "—",
            "Curricula que priorizan":   "; ".join(curricula_mayores) if curricula_mayores else "—",
        })

    if not rows:
        return pd.DataFrame()

    df_rec = pd.DataFrame(rows)
    # Ordenar: primero ALTA urgencia, luego MEDIA, luego subvalorados
    orden_urgencia = {"ALTA": 0, "MEDIA": 1, "—": 2}
    df_rec["_ord"] = df_rec["Urgencia"].map(orden_urgencia)
    df_rec = df_rec.sort_values(["Clasificación", "_ord", "Diferencia (h)"],
                                 ascending=[True, True, False]).drop(columns="_ord")
    return df_rec.reset_index(drop=True)


def export_to_excel(df_rec: pd.DataFrame, df_kpis: dict | None = None) -> bytes:
    """
    Exporta el DataFrame de recomendaciones a Excel con formato de colores.
    Devuelve el contenido como bytes (para st.download_button).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Recomendaciones"

    if df_rec.empty:
        ws["A1"] = "No hay módulos desalineados que reportar."
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    cols = list(df_rec.columns)

    # Encabezados
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill  = FILL_HEADER
        cell.font  = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 30

    # Mapa de urgencia → color de fila
    urgency_fill = {"ALTA": FILL_ALTA, "MEDIA": FILL_MEDIA}

    for row_idx, (_, data_row) in enumerate(df_rec.iterrows(), start=2):
        urgencia = str(data_row.get("Urgencia", "—"))
        clasif   = str(data_row.get("Clasificación", ""))
        fill = urgency_fill.get(urgencia, FILL_BAJO if clasif == "SUBVALORADO" else None)

        for col_idx, col_name in enumerate(cols, start=1):
            val = data_row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if fill:
                # Color suave (opacidad visual parcial)
                alpha_fill = PatternFill("solid", fgColor=_lighten(fill.fgColor.rgb))
                cell.fill = alpha_fill

    # Ajustar anchos de columna automáticamente
    for col_idx, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            *[len(str(df_rec.iloc[r][col_name])) for r in range(len(df_rec))]
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Congelar encabezados
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _lighten(hex_color: str, factor: float = 0.45) -> str:
    """Aclara un color HEX mezclándolo con blanco."""
    hex_color = hex_color.lstrip("#")
    if len(hex_color) == 8:
        hex_color = hex_color[2:]    # quitar alpha si viene como AARRGGBB
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"FF{r:02X}{g:02X}{b:02X}"
