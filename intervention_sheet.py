"""
Genera el Excel 'Intervención_Asignaturas' con 4 secciones (Tarea 4).
"""

from __future__ import annotations
import io
from datetime import date

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from intervention_analysis import INTV_FILL_HEX

# ── Paleta ──────────────────────────────────────────────────────────────────
_H     = PatternFill("solid", fgColor="003366")   # encabezado azul marino
_T_BG  = PatternFill("solid", fgColor="EBF5FB")   # columnas T
_P_BG  = PatternFill("solid", fgColor="EAFAF1")   # columnas P
_SG_BG = PatternFill("solid", fgColor="FEFDE7")   # columnas SG
_ALT   = PatternFill("solid", fgColor="F5F5F5")
_TOT   = PatternFill("solid", fgColor="FFF2CC")
_GOLD  = PatternFill("solid", fgColor="1F2D3D")

_CLASIF_FILL = {k: PatternFill("solid", fgColor=v) for k, v in INTV_FILL_HEX.items()}

_THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)

_SEC_FILLS = {
    1: PatternFill("solid", fgColor="C00000"),
    2: PatternFill("solid", fgColor="1F497D"),
    3: PatternFill("solid", fgColor="2E7D32"),
    4: PatternFill("solid", fgColor="404040"),
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def _c(ws, row, col, val=None, *, bold=False, color="000000", size=10,
       fill=None, ha="center", wrap=False, border=None, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=ha, vertical="center", wrap_text=wrap)
    if fill:   c.fill   = fill
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c


def _sec_title(ws, row, text, fill, nc):
    ws.merge_cells(f"A{row}:{get_column_letter(nc)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = fill
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 24


def _hdr(ws, row, cols, fill=None):
    for ci, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill   = fill or _H
        c.font   = Font(bold=True, color="FFFFFF", size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _THIN
    ws.row_dimensions[row].height = 32


def _col_bg(col_name: str) -> PatternFill | None:
    """Determina el fondo de columna según tipo T/P/SG."""
    low = col_name.lower()
    if "_t" in low or low.endswith(" t") or low == "t":
        return _T_BG
    if "_p" in low or low.endswith(" p") or low == "p":
        return _P_BG
    if "sg" in low:
        return _SG_BG
    return None


def _auto_widths(ws, overrides: dict | None = None):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        if overrides and letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
            continue
        max_len = max((len(str(c.value)) for c in col if c.value), default=6)
        ws.column_dimensions[letter].width = min(max(max_len + 3, 8), 42)


# ── Secciones ────────────────────────────────────────────────────────────────

def _sec1(ws, df_i: pd.DataFrame, cr: int, NC: int) -> int:
    _sec_title(ws, cr,
        f"  SECCIÓN 1 — Ranking por Exceso de Horas  ({len(df_i)} asignaturas)",
        _SEC_FILLS[1], NC)
    cr += 1

    S1 = ["#", "Código", "Asignatura",
          "T", "P", "SG", "Total SHOA",
          "Prom. Internacional", "Diferencia (h)", "Diferencia (%)",
          "% T", "% P", "% SG", "Perfil",
          "Clasificación", "Urgencia", "Tipo hora a reducir"]
    _hdr(ws, cr, S1)
    cr += 1

    for i, (_, r) in enumerate(df_i.iterrows()):
        clf  = r["clasificacion"]
        rfil = _CLASIF_FILL.get(clf)
        alt  = _ALT if i % 2 else None
        base = rfil or alt

        vals_fmt = [
            (i+1,                 None,       "center"),
            (r["codigo_asig"],    None,       "center"),
            (r["nombre_asig"],    None,       "left"),
            (r["shoa_T"],         "0.0",      "center"),
            (r["shoa_P"],         "0.0",      "center"),
            (r["shoa_SG"],        "0.0",      "center"),
            (r["shoa_total"],     "0.0",      "center"),
            (r["intl_avg"],       "0.0",      "center"),
            (r["delta_h"],        "+0.0;-0.0","center"),
            (r["delta_pct"],      "+0.0;-0.0","center"),
            (r["pct_T"],          "0.0",      "center"),
            (r["pct_P"],          "0.0",      "center"),
            (r["pct_SG"],         "0.0",      "center"),
            (r["perfil"],         None,       "center"),
            (clf,                 None,       "center"),
            (clf if clf in ("CRÍTICA","ALTA") else "—", None, "center"),
            (r["tipo_reducir"],   None,       "left"),
        ]

        # Color fills para columnas T/P/SG
        t_fills  = [None, None, None, _T_BG, _P_BG, _SG_BG, None,
                    None, None, None, _T_BG, _P_BG, _SG_BG, None, None, None, None]

        for ci, (val, fmt, ha) in enumerate(vals_fmt, start=1):
            cell_fill = t_fills[ci-1] or base
            c_cell = ws.cell(row=cr, column=ci, value=val)
            c_cell.border    = _THIN
            c_cell.alignment = Alignment(horizontal=ha, vertical="center",
                                         wrap_text=(ci in (3, 17)))
            if cell_fill: c_cell.fill = cell_fill
            if fmt:       c_cell.number_format = fmt
            if ci == 10 and isinstance(val, (int, float)):
                c_cell.font = Font(bold=True,
                                   color="C00000" if val > 0 else "1F497D",
                                   size=10)
        cr += 1

    return cr + 1


def _sec2(ws, df_i: pd.DataFrame, df_exp: pd.DataFrame, cr: int, NC: int) -> int:
    criticas = df_i[df_i["clasificacion"].isin(["CRÍTICA", "ALTA"])]
    _sec_title(ws, cr,
        f"  SECCIÓN 2 — Detalle Tópico a Tópico (asignaturas CRÍTICA/ALTA: {len(criticas)})",
        _SEC_FILLS[2], NC)
    cr += 1

    S2 = ["Asignatura", "Código Tópico", "Descripción Tópico",
          "T", "P", "SG", "Total SHOA",
          "Prom. Internacional", "Diferencia (h)", "Dividida (S/N)"]
    _hdr(ws, cr, S2[:NC] if NC < len(S2) else S2)
    cr += 1

    for _, asig_row in criticas.iterrows():
        code  = asig_row["codigo_asig"]
        topics = df_exp[df_exp["codigo_asig"] == code].copy()

        # Fila de sub-encabezado por asignatura
        ws.merge_cells(f"A{cr}:{get_column_letter(NC)}{cr}")
        c = ws.cell(row=cr, column=1,
                    value=f"  {code} — {asig_row['nombre_asig']}  "
                          f"({asig_row['clasificacion']}, Δ={asig_row['delta_pct']:+.1f}%)")
        c.fill = _CLASIF_FILL.get(asig_row["clasificacion"], _ALT)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        cr += 1

        for j, (_, t) in enumerate(topics.iterrows()):
            alt = _ALT if j % 2 else None
            row_vals = [
                ("", None, "center"),
                (t["codigo_topico"], None, "center"),
                (t["desc_topico"],   None, "left"),
                (t["shoa_T"],        "0.0","center"),
                (t["shoa_P"],        "0.0","center"),
                (t["shoa_SG"],       "0.0","center"),
                (t["shoa_total"],    "0.0","center"),
                (t["intl_avg_topico"],"0.0","center"),
                (round(t["shoa_total"]-t["intl_avg_topico"],1),"+0.0;-0.0","center"),
                ("S" if t["fue_dividida"] else "N", None, "center"),
            ]
            t_fills2 = [None, None, None, _T_BG, _P_BG, _SG_BG,
                        None, None, None, None]
            for ci, (val, fmt, ha) in enumerate(row_vals, start=1):
                c_cell = ws.cell(row=cr, column=ci, value=val)
                c_cell.border    = _THIN
                c_cell.alignment = Alignment(horizontal=ha, vertical="center",
                                             wrap_text=(ci == 3))
                c_cell.fill = t_fills2[ci-1] or alt or PatternFill()
                if fmt: c_cell.number_format = fmt
            cr += 1

    return cr + 1


def _sec3(ws, df_i: pd.DataFrame, cr: int, NC: int) -> int:
    criticas = df_i[df_i["clasificacion"].isin(["CRÍTICA", "ALTA"])]
    _sec_title(ws, cr,
        f"  SECCIÓN 3 — Recomendaciones de Intervención  ({len(criticas)} asignaturas)",
        _SEC_FILLS[3], NC)
    cr += 1

    S3 = ["Código", "Asignatura",
          "T actual", "P actual", "SG actual", "Total actual",
          "T sugerida", "P sugerida", "SG sugerida", "Total sugerido",
          "Reducción T", "Reducción P", "Reducción SG", "Reducción total",
          "Estrategia", "Referencia Internacional"]
    _hdr(ws, cr, S3)
    cr += 1

    for i, (_, r) in enumerate(criticas.iterrows()):
        alt  = _ALT if i % 2 else None
        t_f2 = [None, None, _T_BG, _P_BG, _SG_BG, None,
                _T_BG, _P_BG, _SG_BG, None,
                _T_BG, _P_BG, _SG_BG, None, None, None]
        row_vals = [
            r["codigo_asig"], r["nombre_asig"],
            r["shoa_T"], r["shoa_P"], r["shoa_SG"], r["shoa_total"],
            r["sug_T"],  r["sug_P"],  r["sug_SG"],  r["sug_total"],
            r["red_T"],  r["red_P"],  r["red_SG"],  r["red_total"],
            r["estrategia"], "Promedio internacional equivalente",
        ]
        for ci, val in enumerate(row_vals, start=1):
            c_cell = ws.cell(row=cr, column=ci, value=val)
            c_cell.border    = _THIN
            c_cell.alignment = Alignment(
                horizontal="left" if ci in (2, 15, 16) else "center",
                vertical="center", wrap_text=(ci in (2, 15, 16)))
            c_cell.fill = t_f2[ci-1] or alt or PatternFill()
            if isinstance(val, float): c_cell.number_format = "0.0"
        cr += 1

    return cr + 1


def _sec4(ws, df_i: pd.DataFrame, cr: int, NC: int) -> int:
    _sec_title(ws, cr, "  SECCIÓN 4 — Resumen Ejecutivo", _SEC_FILLS[4], NC)
    cr += 1

    n  = len(df_i)
    c  = df_i["clasificacion"].value_counts().to_dict()
    ca = df_i[df_i["clasificacion"].isin(["CRÍTICA","ALTA"])]

    def pct(x): return f"{round(x/n*100,1)}%" if n else "0%"

    top3 = df_i.nlargest(3, "delta_pct")[["codigo_asig","nombre_asig","delta_pct"]]

    metricas = [
        ("Total asignaturas analizadas",           n),
        ("Asignaturas CRÍTICAS (>30% exceso)",      f"{c.get('CRÍTICA',0)}  ({pct(c.get('CRÍTICA',0))})"),
        ("Asignaturas de intervención ALTA (15–30%)",f"{c.get('ALTA',0)}  ({pct(c.get('ALTA',0))})"),
        ("Asignaturas ALINEADAS (±15%)",            f"{c.get('ALINEADA',0)}  ({pct(c.get('ALINEADA',0))})"),
        ("Asignaturas SUBESTIMADAS (>15% déficit)", f"{c.get('SUBESTIMADA',0)}  ({pct(c.get('SUBESTIMADA',0))})"),
        ("Total horas en exceso — T",               round(ca["red_T"].clip(lower=0).sum(),1)),
        ("Total horas en exceso — P",               round(ca["red_P"].clip(lower=0).sum(),1)),
        ("Total horas en exceso — SG",              round(ca["red_SG"].clip(lower=0).sum(),1)),
        ("Total horas en exceso — General",         round(ca["red_total"].clip(lower=0).sum(),1)),
    ]

    _hdr(ws, cr, ["Métrica", "Valor"])
    cr += 1

    for i, (lbl, val) in enumerate(metricas):
        alt = _ALT if i % 2 else None
        c1 = ws.cell(row=cr, column=1, value=lbl)
        c1.border = _THIN
        c1.alignment = Alignment(horizontal="left", wrap_text=True)
        if alt: c1.fill = alt
        c2 = ws.cell(row=cr, column=2, value=val)
        c2.border = _THIN
        c2.font   = Font(bold=True, size=10)
        c2.alignment = Alignment(horizontal="center")
        if alt: c2.fill = alt
        if isinstance(val, float): c2.number_format = "0.0"
        cr += 1

    cr += 1
    ws.merge_cells(f"A{cr}:B{cr}")
    c = ws.cell(row=cr, column=1, value="  TOP 3 — Asignaturas prioritarias para intervención")
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill("solid", fgColor="7B2222")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[cr].height = 22
    cr += 1

    for j, (_, r) in enumerate(top3.iterrows()):
        ws.cell(row=cr, column=1, value=f"{j+1}. {r['codigo_asig']} — {r['nombre_asig']}").border = _THIN
        ws.cell(row=cr, column=2, value=f"Δ {r['delta_pct']:+.1f}%").border = _THIN
        cr += 1

    return cr


# ── Función principal ────────────────────────────────────────────────────────

def build_intervention_excel(
    df_i: pd.DataFrame,
    df_exp: pd.DataFrame,
) -> bytes:
    """Genera el workbook con las 4 secciones. Retorna bytes para st.download_button."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Intervención_Asignaturas"
    ws.sheet_view.showGridLines = False

    if df_i.empty:
        ws["A1"] = "No hay datos de intervención disponibles."
        buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

    NC = 17   # columnas máximas (sección 1 es la más ancha)
    cr = 1

    # Título global
    ws.merge_cells(f"A1:{get_column_letter(NC)}1")
    c = ws.cell(row=1, column=1,
                value=f"SHOA — Análisis de Intervención por Asignatura T/P/SG  |  {date.today().strftime('%d/%m/%Y')}")
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="003366")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28
    cr = 3

    cr = _sec1(ws, df_i,           cr, NC)
    cr = _sec2(ws, df_i, df_exp,   cr, NC)
    cr = _sec3(ws, df_i,           cr, NC)
    cr = _sec4(ws, df_i,           cr, NC)

    _auto_widths(ws, overrides={"A": 8, "B": 10, "C": 36, "O": 40, "P": 32})
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
